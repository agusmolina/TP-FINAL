#importo las librerias y las clases que se van a usar en el trabajo
import time
import os
from colorama import init,Fore,Style
from openpyxl import Workbook
from openpyxl import load_workbook
from combo import Combo
from registro import Registro
from personas import Encargado, Persona
#Establesco que colorama siempre se reinice al color original luego de usarse para no colorear todo los textos
init(autoreset=True)
#creo una variable de tipo worbook para trabajarla mas facil
wb = Workbook()
#-------------------------------------------------------------------------------------------------------#
#Creo las funciones que voy a necesitar

#valido que el nombre del encargado no este vacioy que no sea un numero, si no lo esta lo pido devuelta
def validarvacio(nombreEmpleado):
   while True:
        if nombreEmpleado==""or nombreEmpleado.isdecimal():
            
            nombreEmpleado=input("Nombre vacio ingrese nuevamente: ")
            
            
        else:
            break
   return nombreEmpleado    

#valido que la variable que me ingresan sea un numero y no este vacia
def convertir(idempleado):
    while True:
        try:
            idempleado = float(idempleado)
            break
        except ValueError:
            print("Numero no valido")
        idempleado = input("Ingrese nuevamente: ")
        print()
    return idempleado        


#Calculo la cuenta del total mutiplicando los combos por su precio
def calculartotal(a,b,c,d):
    total=0
    total=int(comboS.precio)*a+int(comboD.precio)*b+int(comboT.precio)*c+int(comboP.precio)*d
    print("El total a abonar es: "+str(total))
    return total

#calculo el vuelo ademas de sumar las ganancias de la venta a la ganancia general de la empresa
def pagar(total):
    abona=input("Abona con $: ")
    print()
    abona=convertir(abona)
    while True:
        if abona>=total:
            vuelto=abona-total
            print("Su vuelto es de: "+str(vuelto))
            ganancia=total
            mcdowells.gananciadeldia=mcdowells.gananciadeldia+ganancia
            break
        else:
            abona=input("Saldo insuficiente vuelva a pagar: ")
            print()
            abona=convertir(abona)



    
#-------------------------------------------------------------------------------------------------------#

#creo la clase de la empresa
class Mcdowells():
    def __init__(self,nombre="Mc dowell´s",fecha=time.asctime(),listaC=[],ganciadeldia=0):
        self.nombre=nombre
        self.fecha=fecha
        self.lista_combos=listaC
        self.gananciadeldia=ganciadeldia
        

#creo la empresa
mcdowells=Mcdowells()


#Creo los combos usando el constructor de la clase Combo
comboS=Combo("Simple",650,"Hamburguesa simple + Bebida + Fritas")
comboD=Combo("Doble",700,"Hamburguesa doble + Bebida + Fritas")
comboT=Combo("Triple",800,"Hamburguesa Triple + Bebida + Fritas)")
comboP=Combo("Mc Fluby",250,"Helado de dulce de leche")

#Agrego los combos a la lista

mcdowells.lista_combos.append(comboS)
mcdowells.lista_combos.append(comboD)
mcdowells.lista_combos.append(comboT)
mcdowells.lista_combos.append(comboP)

#creo el exel asignandoles a las primeras celdas sus respectivos nombres donde iran guardada la informacion
if os.path.exists(r"C:\Users\Agustin\Desktop\Curso\Trabajo Final\Registro de ventas.xlsx")==False:
    ws = wb.active 
    ws['A1'] = "Cliente"
    ws['B1'] = "fecha"
    ws['c1'] = "Combo S"
    ws['D1'] = "Combo D"
    ws['E1'] = "Combo T"
    ws['F1'] = "Flurby"
    ws['G1'] = "Total"    
    wb.save("Registro de ventas.xlsx")

#-------------------------------------------------------------------------------------------#
#Limpio la pantalla
os.system("cls")

print("Bienvenido a McDowell´s")

nombre=input("Ingrese su nombre encargad@ : "+Fore.RED)
print()
nombre=validarvacio(nombre)

#Creo un nuevo encargado con los datos y otro vacio para hacer el cambio de turno
encargado1=Encargado(nombre,id,)
encargado2=Encargado("",00)

#creo un nuevo tiket donde voy a guardar las salidas y entradas de los encargados
tiket1=Registro("In",time.asctime(),encargado1)

#abro el archivo txt para escribir la entrada del primer empleado del dia
registroventas=open("Registro ventas.txt","a")
n=tiket1.encargado.nombre
n=str(n)
if os.path.exists(r"C:\Users\Agustin\Desktop\Curso\Trabajo Final\Registro ventas.txt")==True:
    registroventas.write(tiket1.accion+" "+tiket1.hora+" "+n.title())
else:
    
    registroventas.write("\n"+tiket1.accion+" "+tiket1.hora+" "+n.title())

registroventas.close()

#--------------------------------------------------------------------------------------#
Fore.RESET
os.system("cls")
#MENU
#creo un boolean para saber si solo un empleado tomo el turno y modificar el guardado del tiket
iterador=False
print("Bienvenido: "+Fore.RED+encargado1.nombre)

while True:
    init(autoreset=True)
    if encargado2.nombre!="":
        os.system("cls")
        print("Bienvenido: "+Fore.RED+encargado2.nombre)

    print(""" 
    McDowell´s
    Recuerda que siempre hay que recibir al cliente con una sonrisa :)
    1 – Ingreso de nuevo pedido
    2 – Cambio de turno
    3 – Apagar sistema
    """)

    opcion=input(">>>")
    opcion=convertir(opcion)

    if opcion==1:
        os.system("cls")
       #pido los datos del cliente
        nomcliente=input("Ingrese el nombre del cliente: ")
        nomcliente=validarvacio(nomcliente)
        apecliente=input("Ingrese el apellido del cliente: ")
        apecliente=validarvacio(apecliente)

        os.system("cls")
        #Agrego cuantos combos compra

        combo1=input("Ingrese cantidad Combo S que quiere "+nomcliente+" "+apecliente+": ")
        combo1=convertir(combo1)
        combo2=input("Ingrese cantidad Combo D que quiere "+nomcliente+" "+apecliente+": ")
        combo2=convertir(combo2)
        combo3=input("Ingrese cantidad Combo T que quiere "+nomcliente+" "+apecliente+": ")
        combo3=convertir(combo3)
        combo4=input("Ingrese cantidad Combo P que quiere "+nomcliente+" "+apecliente+": ")
        combo4=convertir(combo4)

        #calculo el total

        total=calculartotal(combo1,combo2,combo3,combo4)

        #calculo el vuelto
        pagar(total)
        #abro y agrego el registro de la compra al exel
        wb2=load_workbook("Registro de ventas.xlsx")
        ws2=wb2.active
        ws2.append([nomcliente.title()+" "+apecliente.title(),time.asctime(),combo1,combo2,combo3,combo4,total])
        wb2.save("Registro de ventas.xlsx")

        input("---INGRESE ENTER PARA VOLVER AL MENU---")
        os.system("cls")
    elif opcion==2:
        iterador=True
        montoacumulado=mcdowells.gananciadeldia
        os.system("cls")
        print("Bienvenido a McDowell´s")

        #guardo la salida del viejo empleado y la registro en el txt junto al monto acumulado de la opcion 1 
        tiket2=Registro("Out",time.asctime(),encargado1.nombre)
        r=open("Registro ventas.txt","a")
        nom=tiket1.encargado
        nom=str(n)
        r.write("\n"+tiket2.accion+" "+tiket2.hora+" "+nom.title()+" "+str(montoacumulado))
        r.close()
        mcdowells.gananciadeldia=0

        #pregunto los datos del nuevo empleado que ingresa a su turno
        nombre2=input("Ingrese su nombre encargad@ : "+Fore.RED)
        print()
        nombre2=validarvacio(nombre2)
        encargado2.nombre=nombre2

        #registro la entrada del nuevo empleado
        tiket1=Registro("In",time.asctime(),encargado2.nombre)
        nomb=encargado2.nombre
        nomb=str(nomb)
        r=open("Registro ventas.txt","a")
        r.write("\n"+tiket1.accion+" "+tiket1.hora+" "+nomb.title())
        r.close()
    elif opcion==3:
        #cierro el tiket del empleado con la recaudacion de su dia
        os.system("cls")
        montoacumulado=mcdowells.gananciadeldia
        print("¡Gracias por utilizar nuestro programa!")
        if iterador==False :
            nombr=str(encargado1.nombre)
            r=open("Registro ventas.txt","a")
            r.write("\n"+"out"+" "+time.asctime()+" "+nombr.title()+" "+str(montoacumulado))
            r.close()
        else:
            r=open("Registro ventas.txt","a")
            r.write("\n"+tiket2.accion+" "+time.asctime()+" "+nomb.title()+" "+str(montoacumulado))
            r.close()
        break
    else:
        print("¡OPCION NO VALIDA!")
        time.sleep(1)
        os.system("cls")